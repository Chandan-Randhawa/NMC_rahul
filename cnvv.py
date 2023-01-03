from openpyxl import load_workbook
import csv

wb = load_workbook(filename= 'NMC data v2.xlsx')
sheet = wb.active

csv_data=[]
for value in sheet.iter_rows(values_only=True):
    csv_data.append(list(value))

with open('NMC data v2.csv', 'w') as csv_obj:
    writer =csv.writer(csv_obj , delimiter=',')
    for line in csv_data:
        writer.writerow(line)